#!/usr/bin/env python3
"""Profile release asset contents without committing the source payloads.

The profiler downloads each unique release payload to temporary storage and
records only structural signals (headings, sheet names, keyword counts, file
shape). It deliberately does not copy full release payloads into the repo.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
import urllib.error
import urllib.request
import zipfile
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MANIFEST = REPO_ROOT / "raw/release-assets/backend-docs/MANIFEST.json"
DEFAULT_OUTPUT = REPO_ROOT / "raw/release-assets/backend-docs/CONTENT_INDEX.json"
DEFAULT_REPORT = REPO_ROOT / "docs/RELEASE_ASSET_CONTENT_PROFILE.md"

KEYWORDS = (
    "sourceborn", "urr", "asi", "engine", "definition", "parameter",
    "sequence", "wisdom", "human", "brain", "node", "registry",
    "rubric", "verification", "audit", "test", "grok", "riemann",
    "mirror", "render", "deploy", "stock", "memory", "feedback",
)

TEXT_EXTENSIONS = {".md", ".txt", ".csv", ".json", ".jsonl", ".yaml", ".yml", ".xml", ".tsv", ".log"}


def auth_headers(token: str, octet_stream: bool = False) -> dict[str, str]:
    headers = {
        "Accept": "application/octet-stream" if octet_stream else "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "c-sb-release-asset-profiler",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def download(url: str, token: str, path: Path) -> None:
    req = urllib.request.Request(url, headers=auth_headers(token, octet_stream=True))
    try:
        with urllib.request.urlopen(req, timeout=120) as response, path.open("wb") as out:
            while True:
                block = response.read(1024 * 1024)
                if not block:
                    break
                out.write(block)
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"download failed {exc.code} for {url}: {body[:500]}") from exc


def clean_line(value: str, limit: int = 240) -> str:
    value = re.sub(r"\s+", " ", value).strip()
    return value[:limit]


def keyword_counts(text: str) -> dict[str, int]:
    lower = text.lower()
    return {k: lower.count(k) for k in KEYWORDS if lower.count(k)}


def profile_text(path: Path) -> dict[str, Any]:
    data = path.read_bytes()
    text = data.decode("utf-8", errors="replace")
    lines = text.splitlines()
    nonempty = [clean_line(line) for line in lines if line.strip()]
    headings = []
    for line in lines:
        stripped = line.strip()
        if re.match(r"^#{1,6}\s+", stripped):
            headings.append(clean_line(stripped))
        if len(headings) >= 30:
            break
    result: dict[str, Any] = {
        "kind": "text",
        "line_count": len(lines),
        "character_count": len(text),
        "headings": headings,
        "opening_lines": nonempty[:10],
        "keyword_hits": keyword_counts(text),
    }
    if path.suffix.lower() == ".json":
        try:
            parsed = json.loads(text)
            result["json_shape"] = type(parsed).__name__
            if isinstance(parsed, dict):
                result["json_top_level_keys"] = list(parsed.keys())[:40]
            elif isinstance(parsed, list):
                result["json_list_length"] = len(parsed)
        except json.JSONDecodeError:
            result["json_parse_error"] = True
    if path.suffix.lower() in {".csv", ".tsv"}:
        sep = "\t" if path.suffix.lower() == ".tsv" else ","
        if nonempty:
            result["table_header"] = [clean_line(x, 100) for x in nonempty[0].split(sep)[:40]]
    return result


def profile_xlsx(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to profile xlsx assets") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    aggregate = Counter()
    for ws in wb.worksheets:
        first_rows: list[list[str]] = []
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            first_rows.append([clean_line(str(v), 120) if v is not None else "" for v in row[:30]])
        sample_text = " ".join(cell for row in first_rows for cell in row if cell)
        aggregate.update(keyword_counts(sample_text))
        sheets.append({
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "first_rows": first_rows,
        })
    wb.close()
    return {
        "kind": "xlsx",
        "sheet_count": len(sheets),
        "sheets": sheets,
        "keyword_hits_in_headers": dict(aggregate),
    }


def profile_docx(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path) as zf:
        raw = zf.read("word/document.xml")
    root = ET.fromstring(raw)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs = []
    for p in root.findall(".//w:p", ns):
        pieces = [t.text or "" for t in p.findall(".//w:t", ns)]
        text = clean_line("".join(pieces))
        if text:
            paragraphs.append(text)
    whole = "\n".join(paragraphs)
    return {
        "kind": "docx",
        "paragraph_count": len(paragraphs),
        "opening_paragraphs": paragraphs[:20],
        "keyword_hits": keyword_counts(whole),
    }


def profile_pdf(path: Path) -> dict[str, Any]:
    try:
        from pypdf import PdfReader
    except ImportError:
        return {"kind": "pdf", "note": "pypdf unavailable"}
    reader = PdfReader(str(path))
    snippets = []
    whole_parts = []
    for page in reader.pages[:20]:
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        whole_parts.append(text)
        for line in text.splitlines():
            line = clean_line(line)
            if line:
                snippets.append(line)
            if len(snippets) >= 20:
                break
        if len(snippets) >= 20:
            break
    whole = "\n".join(whole_parts)
    return {
        "kind": "pdf",
        "page_count": len(reader.pages),
        "opening_lines": snippets[:20],
        "keyword_hits_sample": keyword_counts(whole),
    }


def suggested_route(asset: dict[str, Any], profile: dict[str, Any]) -> tuple[str, str]:
    current = asset.get("route", "unclassified")
    if current not in {"unclassified"}:
        return current, "filename route retained"

    blob = json.dumps(profile, ensure_ascii=False).lower()
    hits = Counter()
    for k in KEYWORDS:
        hits[k] = blob.count(k)

    if hits["grok"]:
        return "quarantine", "content signal contains Grok"
    if hits["stock"] >= 2:
        return "off-project", "content signal is stock-related"
    if hits["riemann"] or hits["mirror"] >= 2:
        return "research", "content signal is RH/Riemann/Mirror research"
    if hits["sourceborn"] >= 3 or hits["urr"] >= 3:
        return "sourceborn-core", "strong Sourceborn/URR content signal"
    if hits["asi"] >= 4 and (hits["brain"] or hits["node"] or hits["registry"]):
        return "asi", "strong ASI architecture/registry signal"
    if hits["engine"] >= 4 or hits["definition"] >= 3:
        return "engines", "strong engine/definition content signal"
    if hits["parameter"] >= 4:
        return "parameters", "strong parameter content signal"
    if hits["sequence"] >= 4:
        return "sequence", "strong sequence content signal"
    if hits["wisdom"] >= 3:
        return "wisdom", "strong wisdom content signal"
    if hits["audit"] + hits["test"] + hits["verification"] >= 5:
        return "tests-audits", "strong testing/audit content signal"

    name = asset.get("name", "").lower()
    if any(x in name for x in ("conversation", "transcript", "chat", "archive", "memory", "pasted")):
        return "archives-transcripts", "conversation/archive filename signal"
    if any(x in name for x in ("claude", "gemini", "gpt", "lovable", "model_")):
        return "model-evidence", "model-output/reference filename signal"
    if any(x in name for x in ("ard", "rgl", "met_")):
        return "research", "ARD/RGL research/reference filename signal"
    return "reference", "generic source/reference material pending semantic promotion"


def profile_one(asset: dict[str, Any], token: str, tmpdir: Path) -> dict[str, Any]:
    suffix = Path(asset["name"]).suffix.lower()
    path = tmpdir / f"{asset['asset_id']}{suffix or '.bin'}"
    result: dict[str, Any] = {
        "asset_id": asset["asset_id"],
        "name": asset["name"],
        "sha256": asset.get("sha256"),
        "size": asset.get("size"),
        "manifest_route": asset.get("route"),
    }
    try:
        download(asset["api_url"], token, path)
        if suffix in TEXT_EXTENSIONS:
            profile = profile_text(path)
        elif suffix == ".xlsx":
            profile = profile_xlsx(path)
        elif suffix == ".docx":
            profile = profile_docx(path)
        elif suffix == ".pdf":
            profile = profile_pdf(path)
        elif zipfile.is_zipfile(path):
            with zipfile.ZipFile(path) as zf:
                profile = {"kind": "zip", "entries": zf.namelist()[:100], "entry_count": len(zf.namelist())}
        else:
            profile = {"kind": "binary"}
        route, reason = suggested_route(asset, profile)
        result.update({"profile": profile, "suggested_route": route, "suggested_route_reason": reason, "status": "profiled"})
    except Exception as exc:
        result.update({"status": "error", "error": str(exc), "suggested_route": asset.get("route"), "suggested_route_reason": "profiling failed; retained manifest route"})
    finally:
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass
    return result


def write_report(index: dict[str, Any], path: Path) -> None:
    profiles = index["profiles"]
    suggested_counts = Counter(p.get("suggested_route", "unknown") for p in profiles)
    errors = [p for p in profiles if p.get("status") == "error"]
    changed = [p for p in profiles if p.get("suggested_route") != p.get("manifest_route")]

    lines = [
        "# Release Asset Content Profile",
        "",
        "This report profiles **unique** release payloads structurally. Full payloads remain in the GitHub release and are not copied into the repository.",
        "",
        f"- Unique payloads profiled: **{len(profiles)}**",
        f"- Profiling errors: **{len(errors)}**",
        f"- Routes refined by content/fallback signals: **{len(changed)}**",
        "",
        "## Suggested routing after content profiling",
        "",
        "| Route | Unique payloads |",
        "|---|---:|",
    ]
    for route, count in suggested_counts.most_common():
        lines.append(f"| `{route}` | {count} |")

    lines += ["", "## Reclassified unique payloads", "", "| From | To | Asset | Reason |", "|---|---|---|---|"]
    for p in sorted(changed, key=lambda x: (x.get("suggested_route", ""), x.get("name", "").lower())):
        reason = str(p.get("suggested_route_reason", "")).replace("|", "\\|")
        name = str(p.get("name", "")).replace("|", "\\|")
        lines.append(f"| `{p.get('manifest_route')}` | `{p.get('suggested_route')}` | `{name}` | {reason} |")

    if errors:
        lines += ["", "## Profiling errors", ""]
        for p in errors:
            lines.append(f"- `{p['name']}`: {p.get('error')}")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--workers", type=int, default=8)
    args = parser.parse_args()

    token = os.environ.get("GITHUB_TOKEN", "")
    manifest = json.loads(args.manifest.read_text(encoding="utf-8"))
    unique_assets = [a for a in manifest["assets"] if a.get("digest_primary")]

    profiles = []
    with tempfile.TemporaryDirectory(prefix="csb-release-profile-") as td:
        tmpdir = Path(td)
        with ThreadPoolExecutor(max_workers=max(1, args.workers)) as pool:
            futures = {pool.submit(profile_one, asset, token, tmpdir): asset for asset in unique_assets}
            for future in as_completed(futures):
                profile = future.result()
                profiles.append(profile)
                print(f"{profile['status']:8} {profile['name']}")

    profiles.sort(key=lambda p: p["name"].lower())
    index = {
        "schema_version": 1,
        "release_id": manifest["release"]["id"],
        "release_tag": manifest["release"]["tag_name"],
        "unique_payload_count": len(unique_assets),
        "profiles": profiles,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(index, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    write_report(index, args.report)

    errors = sum(1 for p in profiles if p["status"] == "error")
    print(json.dumps({"profiled": len(profiles), "errors": errors}, indent=2))
    if len(profiles) != manifest["summary"]["unique_sha256_payloads"]:
        raise RuntimeError("profile count does not match unique payload count")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
