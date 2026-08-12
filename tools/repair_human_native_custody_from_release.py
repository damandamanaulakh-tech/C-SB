#!/usr/bin/env python3
"""Rebuild the missing/corrupt Human-2560 custody parts from release source.

This is intentionally narrow and fail-closed.  It accepts only the approved
release workbook/digest and only the exact first 2,560 canonical Brain Base
parameter rows required by materialize_human_native_2560_v1.py.
"""

from __future__ import annotations

import argparse
import base64
import csv
import gzip
import hashlib
import io
import json
import os
import re
import tempfile
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MANIFEST = ROOT / "raw/release-assets/backend-docs/MANIFEST.json"
OUT_DIR = ROOT / "raw/human/task3_human_native_parts"
SOURCE_MANIFEST = OUT_DIR / "SOURCE_MANIFEST.json"

SOURCE_NAME = "ASI_Brain_Engine_Combined_Corpus_v1.xlsx"
SOURCE_ASSET_ID = 512182897
SOURCE_SHA256 = "6d6bd608844b07728aaefb0d16e6c36bfcf7ba4ac3ec70af2610ea2bd7622a1b"
SOURCE_SHEET = "03 Existing Parameters"
CHUNK_ROWS = 320
EXPECTED_RECORDS = 2560
EXPECTED_PARTS = 8

PID_RE = re.compile(r"^SB-ASI-P(\d{4})$")
CON_RE = re.compile(r"^CON-(\d{3})$")
SEG_RE = re.compile(r"^SEG-(\d{2})$")
EXPECTED_IDS = [f"SB-ASI-P{i:04d}" for i in range(1, EXPECTED_RECORDS + 1)]


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def download(url: str, token: str, path: Path) -> None:
    headers = {
        "Accept": "application/octet-stream",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "c-sb-human-custody-repair",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=120) as response, path.open("wb") as out:
        while True:
            block = response.read(1024 * 1024)
            if not block:
                break
            out.write(block)


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def trim(values: list[str]) -> list[str]:
    result = list(values)
    while result and result[-1] == "":
        result.pop()
    return result


def load_source_asset() -> dict[str, Any]:
    data = json.loads(MANIFEST.read_text(encoding="utf-8"))
    matches = [a for a in data["assets"] if a.get("name") == SOURCE_NAME]
    if len(matches) != 1:
        raise RuntimeError(f"expected one manifest record for {SOURCE_NAME}, found {len(matches)}")
    asset = matches[0]
    if int(asset["asset_id"]) != SOURCE_ASSET_ID:
        raise RuntimeError(f"source asset id drift: {asset['asset_id']} != {SOURCE_ASSET_ID}")
    if asset.get("sha256") != SOURCE_SHA256:
        raise RuntimeError(f"source manifest digest drift: {asset.get('sha256')} != {SOURCE_SHA256}")
    return asset


def extract_rows(workbook_path: Path) -> tuple[list[str], list[list[str]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if SOURCE_SHEET not in wb.sheetnames:
        raise RuntimeError(f"required sheet missing: {SOURCE_SHEET}")
    ws = wb[SOURCE_SHEET]

    header: list[str] | None = None
    records: list[list[str]] = []
    ids: list[str] = []

    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = trim([cell_text(v) for v in row])
        if row_index == 1:
            header = values
            continue
        pids = [v for v in values if PID_RE.fullmatch(v)]
        if not pids:
            continue
        pid = pids[0]
        number = int(PID_RE.fullmatch(pid).group(1))
        if number > EXPECTED_RECORDS:
            continue
        records.append(values)
        ids.append(pid)

    wb.close()
    if header is None:
        raise RuntimeError("source sheet has no header")
    if ids != EXPECTED_IDS:
        missing = [pid for pid in EXPECTED_IDS if pid not in set(ids)]
        raise RuntimeError(
            f"source ID order/coverage mismatch count={len(ids)} first={ids[0] if ids else None} "
            f"last={ids[-1] if ids else None} missing={missing[:10]}"
        )
    if len(header) != 13:
        raise RuntimeError(f"source header columns != 13: {len(header)} {header}")
    bad_columns = [(idx + 1, len(row)) for idx, row in enumerate(records) if len(row) != 13]
    if bad_columns:
        raise RuntimeError(f"parameter rows with non-13 column shape: {bad_columns[:10]}")

    containers = {v for row in records for v in row if CON_RE.fullmatch(v)}
    segments = {v for row in records for v in row if SEG_RE.fullmatch(v)}
    if len(containers) != 80:
        raise RuntimeError(f"selected source rows contain {len(containers)} containers, expected 80")
    if len(segments) != 10:
        raise RuntimeError(f"selected source rows contain {len(segments)} segments, expected 10")

    approval = sum("APPROVED BY USER" in row for row in records)
    evident = sum("USER EVIDENT" in row for row in records)
    brain_base = sum("Canonical Brain Base" in row for row in records)
    if (approval, evident, brain_base) != (EXPECTED_RECORDS, EXPECTED_RECORDS, EXPECTED_RECORDS):
        raise RuntimeError(
            f"source marker coverage mismatch approval={approval} evident={evident} brain_base={brain_base}"
        )
    return header, records


def rows_to_tsv(rows: list[list[str]]) -> str:
    stream = io.StringIO(newline="")
    writer = csv.writer(stream, delimiter="\t", lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
    writer.writerows(rows)
    return stream.getvalue()


def write_parts(header: list[str], records: list[list[str]], source: dict[str, Any]) -> dict[str, Any]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    full_tsv = rows_to_tsv([header] + records)
    part_records: list[dict[str, Any]] = []
    reconstructed_chunks: list[str] = []

    for part_number in range(1, EXPECTED_PARTS + 1):
        start = (part_number - 1) * CHUNK_ROWS
        end = start + CHUNK_ROWS
        chunk_records = records[start:end]
        if len(chunk_records) != CHUNK_ROWS:
            raise RuntimeError(f"part {part_number} record count {len(chunk_records)} != {CHUNK_ROWS}")
        rows = ([header] if part_number == 1 else []) + chunk_records
        tsv = rows_to_tsv(rows)
        reconstructed_chunks.append(tsv)
        gz = gzip.compress(tsv.encode("utf-8"), compresslevel=9, mtime=0)
        encoded = base64.b64encode(gz).decode("ascii") + "\n"
        path = OUT_DIR / f"part-{part_number:02d}.tsv.gz.b64"
        path.write_text(encoded, encoding="ascii")

        # Fail immediately if our own transport cannot round-trip independently.
        decoded = gzip.decompress(base64.b64decode(encoded, validate=False)).decode("utf-8")
        if decoded != tsv:
            raise RuntimeError(f"part {part_number} failed deterministic transport round-trip")
        part_records.append(
            {
                "part": part_number,
                "path": str(path.relative_to(ROOT)),
                "parameter_start": EXPECTED_IDS[start],
                "parameter_end": EXPECTED_IDS[end - 1],
                "parameter_rows": len(chunk_records),
                "includes_header": part_number == 1,
                "tsv_sha256": sha256_bytes(tsv.encode("utf-8")),
                "gzip_sha256": sha256_bytes(gz),
                "file_sha256": sha256_file(path),
                "file_size": path.stat().st_size,
            }
        )

    expected_paths = {OUT_DIR / f"part-{i:02d}.tsv.gz.b64" for i in range(1, EXPECTED_PARTS + 1)}
    for path in OUT_DIR.glob("part-*.tsv.gz.b64"):
        if path not in expected_paths:
            path.unlink()

    concatenated = "".join(reconstructed_chunks)
    if concatenated != full_tsv:
        raise RuntimeError("concatenated independent custody chunks do not reproduce full source TSV")

    source_manifest = {
        "schema_version": 1,
        "status": "REBUILT_FROM_APPROVED_RELEASE_SOURCE",
        "source": {
            "release_tag": "Docs",
            "release_id": 369575518,
            "asset_id": source["asset_id"],
            "asset_name": source["name"],
            "asset_sha256": source["sha256"],
            "sheet": SOURCE_SHEET,
            "selection": "header + ordered parameter rows SB-ASI-P0001..SB-ASI-P2560 only",
        },
        "shape": {
            "parameter_rows": EXPECTED_RECORDS,
            "part_count": EXPECTED_PARTS,
            "parameter_rows_per_part": CHUNK_ROWS,
            "columns_per_parameter_row": 13,
            "container_count": 80,
            "segment_count": 10,
        },
        "reconstructed_tsv_sha256": sha256_bytes(full_tsv.encode("utf-8")),
        "parts": part_records,
        "validation_rule": "tools/materialize_human_native_2560_v1.py must pass before custody is committed",
    }
    SOURCE_MANIFEST.write_text(json.dumps(source_manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return source_manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.parse_args()

    source = load_source_asset()
    token = os.environ.get("GITHUB_TOKEN", "")
    with tempfile.TemporaryDirectory(prefix="csb-human-custody-repair-") as td:
        workbook = Path(td) / SOURCE_NAME
        download(source["api_url"], token, workbook)
        actual = sha256_file(workbook)
        if actual != SOURCE_SHA256:
            raise RuntimeError(f"downloaded source SHA-256 mismatch: {actual} != {SOURCE_SHA256}")
        header, records = extract_rows(workbook)
        source_manifest = write_parts(header, records, source)

    print(json.dumps({
        "status": source_manifest["status"],
        "source": source_manifest["source"],
        "shape": source_manifest["shape"],
        "reconstructed_tsv_sha256": source_manifest["reconstructed_tsv_sha256"],
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
