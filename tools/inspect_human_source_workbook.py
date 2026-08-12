#!/usr/bin/env python3
"""Inspect candidate Human-2560 release workbooks without committing binaries.

The existing Human custody materializer defines the authoritative transport
shape: SB-ASI-P0001..SB-ASI-P2560, 80 CON-xxx containers, 10 SEG-xx segments,
and 13 columns per parameter row. This inspector looks for that exact source
shape instead of inventing a new Human ID namespace.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
import urllib.request
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MANIFEST = ROOT / "raw/release-assets/backend-docs/MANIFEST.json"
DEFAULT_OUTPUT = ROOT / "raw/release-assets/backend-docs/HUMAN_SOURCE_INSPECTION.json"
DEFAULT_REPORT = ROOT / "docs/HUMAN_SOURCE_WORKBOOK_INSPECTION.md"

CANDIDATE_NAMES = (
    "ASI_Brain_Engine_Combined_Corpus_v1.xlsx",
    "Brain.+.Engine.Combined.Corpus.xlsx",
    "ASI-Brain.xlsx",
    "ASI-Brain_Task2_Approved_v0_1.xlsx",
    "ASI-Brain_Core_Engine_Combined_v0_4.xlsx",
    "ASI-Brain_Merged_APPROVED_EVIDENT_v0_3.xlsx",
    "ASI-Brain_Task3_Review_v0_2.xlsx",
)

PID_RE = re.compile(r"^SB-ASI-P(\d{4})$")
CON_RE = re.compile(r"^CON-(\d{3})$")
SEG_RE = re.compile(r"^SEG-(\d{2})$")
EXPECTED_IDS = [f"SB-ASI-P{i:04d}" for i in range(1, 2561)]
EXPECTED_ID_SET = set(EXPECTED_IDS)
SENTINELS = {
    "SB-ASI-P0001", "SB-ASI-P0032", "SB-ASI-P0033", "SB-ASI-P0256",
    "SB-ASI-P0257", "SB-ASI-P0512", "SB-ASI-P0513", "SB-ASI-P1280",
    "SB-ASI-P1281", "SB-ASI-P1920", "SB-ASI-P1921", "SB-ASI-P2560",
}


def download(url: str, token: str, path: Path) -> None:
    headers = {
        "Accept": "application/octet-stream",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "c-sb-human-source-inspector",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=120) as response, path.open("wb") as out:
        while True:
            block = response.read(1024 * 1024)
            if not block:
                break
            out.write(block)


def norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def trim_row(values: list[str]) -> list[str]:
    values = list(values)
    while values and values[-1] == "":
        values.pop()
    return values


def inspect_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, Any] = {"sheets": []}
    workbook_ids: set[str] = set()
    workbook_containers: set[str] = set()
    workbook_segments: set[str] = set()

    for ws in wb.worksheets:
        ids: list[str] = []
        containers: set[str] = set()
        segments: set[str] = set()
        param_column_counts: Counter[int] = Counter()
        approval_count = 0
        evidence_count = 0
        brain_base_count = 0
        first_parameter_rows: list[dict[str, Any]] = []
        sentinel_rows: dict[str, dict[str, Any]] = {}
        first_rows: list[list[str]] = []
        scanned_rows = 0
        max_seen_columns = 0

        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            scanned_rows = r_idx
            vals = trim_row([norm(v) for v in row])
            max_seen_columns = max(max_seen_columns, len(vals))
            if r_idx <= 8:
                first_rows.append(vals[:30])

            row_pids = [v for v in vals if PID_RE.fullmatch(v)]
            if not row_pids:
                continue
            pid = row_pids[0]
            ids.append(pid)
            workbook_ids.add(pid)
            row_cons = {v for v in vals if CON_RE.fullmatch(v)}
            row_segs = {v for v in vals if SEG_RE.fullmatch(v)}
            containers.update(row_cons)
            segments.update(row_segs)
            workbook_containers.update(row_cons)
            workbook_segments.update(row_segs)
            param_column_counts[len(vals)] += 1
            approval_count += int("APPROVED BY USER" in vals)
            evidence_count += int("USER EVIDENT" in vals)
            brain_base_count += int("Canonical Brain Base" in vals)

            if len(first_parameter_rows) < 5:
                first_parameter_rows.append({"row": r_idx, "values": vals[:30]})
            if pid in SENTINELS and pid not in sentinel_rows:
                sentinel_rows[pid] = {"row": r_idx, "values": vals[:30]}

        id_set = set(ids)
        sheet = {
            "name": ws.title,
            "scanned_rows": scanned_rows,
            "max_seen_columns": max_seen_columns,
            "parameter_row_count": len(ids),
            "unique_parameter_id_count": len(id_set),
            "first_parameter_id": ids[0] if ids else None,
            "last_parameter_id": ids[-1] if ids else None,
            "ordered_id_coverage_exact_0001_2560": ids == EXPECTED_IDS,
            "set_id_coverage_exact_0001_2560": id_set == EXPECTED_ID_SET,
            "missing_parameter_ids": [pid for pid in EXPECTED_IDS if pid not in id_set][:25],
            "container_count": len(containers),
            "segment_count": len(segments),
            "parameter_column_counts": dict(sorted(param_column_counts.items())),
            "approval_coverage": approval_count,
            "evidence_coverage": evidence_count,
            "canonical_brain_base_coverage": brain_base_count,
            "first_rows": first_rows,
            "first_parameter_rows": first_parameter_rows,
            "sentinel_rows": sentinel_rows,
        }
        result["sheets"].append(sheet)

    wb.close()
    result.update({
        "workbook_unique_parameter_ids": len(workbook_ids),
        "workbook_parameter_id_set_exact_0001_2560": workbook_ids == EXPECTED_ID_SET,
        "workbook_container_count": len(workbook_containers),
        "workbook_segment_count": len(workbook_segments),
    })
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    args = parser.parse_args()

    manifest = json.loads(args.manifest.read_text(encoding="utf-8"))
    by_name = {a["name"]: a for a in manifest["assets"]}
    token = os.environ.get("GITHUB_TOKEN", "")
    inspections = []

    with tempfile.TemporaryDirectory(prefix="csb-human-source-") as td:
        tmp = Path(td)
        for name in CANDIDATE_NAMES:
            asset = by_name.get(name)
            if not asset:
                continue
            path = tmp / f"{asset['asset_id']}.xlsx"
            print(f"Inspecting {name}...")
            download(asset["api_url"], token, path)
            inspected = inspect_workbook(path)
            inspected.update({
                "asset_id": asset["asset_id"],
                "name": name,
                "sha256": asset.get("sha256"),
                "size": asset.get("size"),
            })
            inspections.append(inspected)

    payload = {"schema_version": 2, "candidates": inspections}
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    lines = [
        "# Human-2560 Source Workbook Inspection",
        "",
        "Release binaries are downloaded only inside CI and are not committed. The inspection uses the exact identifiers and row-shape enforced by `tools/materialize_human_native_2560_v1.py`.",
        "",
        "Required source shape: `SB-ASI-P0001..SB-ASI-P2560`, 80 `CON-xxx` containers, 10 `SEG-xx` segments, 13 columns per parameter row, with the approval/evidence/base markers expected by the materializer.",
        "",
    ]
    for item in inspections:
        lines += [
            f"## `{item['name']}`",
            "",
            f"- Asset: `{item['asset_id']}`",
            f"- SHA-256: `{item['sha256']}`",
            f"- Size: `{item['size']}` bytes",
            f"- Unique `SB-ASI-P` IDs across workbook: **{item['workbook_unique_parameter_ids']}**",
            f"- Exact ID set P0001..P2560 present: **{item['workbook_parameter_id_set_exact_0001_2560']}**",
            f"- Containers seen: **{item['workbook_container_count']}**",
            f"- Segments seen: **{item['workbook_segment_count']}**",
            "",
            "| Sheet | Rows scanned | Parameter rows | IDs | First → Last | Ordered full set | Containers | Segments | Column counts | Approval | Evident | Brain Base |",
            "|---|---:|---:|---:|---|---|---:|---:|---|---:|---:|---:|",
        ]
        for sheet in item["sheets"]:
            if not sheet["parameter_row_count"]:
                continue
            col_counts = ", ".join(f"{k}:{v}" for k, v in sheet["parameter_column_counts"].items())
            lines.append(
                f"| `{sheet['name']}` | {sheet['scanned_rows']} | {sheet['parameter_row_count']} | {sheet['unique_parameter_id_count']} | "
                f"`{sheet['first_parameter_id']}` → `{sheet['last_parameter_id']}` | {sheet['ordered_id_coverage_exact_0001_2560']} | "
                f"{sheet['container_count']} | {sheet['segment_count']} | `{col_counts}` | {sheet['approval_coverage']} | "
                f"{sheet['evidence_coverage']} | {sheet['canonical_brain_base_coverage']} |"
            )
        lines += ["", "### Matching parameter row samples", ""]
        for sheet in item["sheets"]:
            if not sheet["parameter_row_count"]:
                continue
            lines.append(f"**{sheet['name']}**")
            for sample in sheet["first_parameter_rows"]:
                compact = " | ".join(sample["values"])
                lines.append(f"- row {sample['row']}: `{compact[:1800]}`")
            if sheet["sentinel_rows"]:
                lines.append("- sentinel IDs located: " + ", ".join(sorted(sheet["sentinel_rows"])))
            if sheet["missing_parameter_ids"]:
                lines.append("- first missing IDs: " + ", ".join(sheet["missing_parameter_ids"]))
            lines.append("")

    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text("\n".join(lines), encoding="utf-8")

    candidates = [
        {
            "name": c["name"],
            "parameter_ids": c["workbook_unique_parameter_ids"],
            "full_2560": c["workbook_parameter_id_set_exact_0001_2560"],
            "containers": c["workbook_container_count"],
            "segments": c["workbook_segment_count"],
        }
        for c in inspections
    ]
    print(json.dumps(candidates, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
